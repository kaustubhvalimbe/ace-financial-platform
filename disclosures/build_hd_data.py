#!/usr/bin/env python3
"""
Turns a folder of AMC monthly portfolio disclosure spreadsheets into hd-data.json
for stock-intelligence.html / portfolio-xray.html.

MONTHLY WORKFLOW
----------------
1. Go to https://www.amfiindia.com/online-center/portfolio-disclosure
   Download each AMC's monthly portfolio disclosure spreadsheet (xlsx or csv;
   all AMCs publish in the same SEBI-mandated format: ISIN, Name of Instrument,
   Industry/Rating, Quantity, Market Value, % to NAV -- one block per scheme).
2. Save each file into disclosures/raw/<YYYY-MM>/ using the AMC's name as the
   filename, e.g. disclosures/raw/2026-07/HDFC MF.xlsx
   (underscores in filenames are treated as spaces, e.g. HDFC_MF.xlsx also works)
3. Run:
     python disclosures/build_hd_data.py 2026-07
   This writes hd-data.json in the site root (after backing up the old one)
   and disclosures/review/<YYYY-MM>-unmapped-*.csv for anything it couldn't
   confidently classify.
4. Open the unmapped-*.csv files. For any stock/sector listed, add an entry
   to disclosures/lookups/stock_symbols.json or sector_map.json (a 2-minute
   job the first time a new stock/label shows up -- it then stays fixed for
   every future month).
5. Re-run the same command to pick up your lookup additions, sanity-check
   the printed summary, then deploy as usual (deploy.sh / deploy.bat).

Requires: pip install openpyxl
"""
import csv
import datetime
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DISCLOSURES = ROOT / "disclosures"
LOOKUPS = DISCLOSURES / "lookups"
REVIEW = DISCLOSURES / "review"
HD_DATA = ROOT / "hd-data.json"

# Column header aliases (lowercased substring match)
COL_ALIASES = {
    "isin": ["isin"],
    "name": ["name of instrument", "name of the instrument", "instrument name", "company name"],
    "industry": ["industry", "industry / rating", "industry+/ rating", "sector"],
    "quantity": ["quantity", "qty"],
    "value": ["market value", "market value(rs", "market value (rs"],
    "pct": ["% to nav", "% to net assets", "%age to nav", "percentage to nav", "% to net asset"],
}

SCHEME_MARKER = re.compile(r"name of (the )?(scheme|fund)\s*:?\s*", re.I)

# Instrument categories that are never individual equity stocks -- skip these rows
NON_EQUITY_HINTS = [
    "treps", "cash", "net current assets", "cblo", "government securit",
    "certificate of deposit", "commercial paper", "corporate debt",
    "reverse repo", "margin", "net receivable", "net payable",
    "money market", "mutual fund unit", "clearing corporation",
]

CORP_SUFFIX_RE = re.compile(
    r"\s+(limited|ltd\.?|the\s*$)\s*$", re.I
)


def clean_stock_name(raw: str) -> str:
    name = raw.strip()
    name = CORP_SUFFIX_RE.sub("", name).strip()
    name = re.sub(r"\s+", " ", name)
    return name


def find_header_row(rows, start_idx, max_scan=25):
    """Scan forward from start_idx for a row that looks like the column header."""
    for i in range(start_idx, min(start_idx + max_scan, len(rows))):
        cells = [str(c).strip().lower() if c is not None else "" for c in rows[i]]
        has_isin = any(any(a in c for a in COL_ALIASES["isin"]) for c in cells)
        has_pct = any(any(a in c for a in COL_ALIASES["pct"]) for c in cells)
        if has_isin and has_pct:
            return i, cells
    return None, None


def map_columns(header_cells):
    col_idx = {}
    for key, aliases in COL_ALIASES.items():
        for i, cell in enumerate(header_cells):
            if any(a in cell for a in aliases):
                col_idx[key] = i
                break
    return col_idx


def is_equity_row(isin, name, industry):
    if not isin or not name:
        return False
    isin = str(isin).strip().upper()
    if not isin.startswith("INE"):
        return False
    blob = f"{name} {industry}".lower()
    if any(h in blob for h in NON_EQUITY_HINTS):
        return False
    return True


def parse_sheet(rows, sheet_name):
    """Yields (scheme_name, isin, name, industry, pct) for every equity holding row
    found anywhere in the sheet, splitting on 'Name of the Scheme' markers when present."""
    n = len(rows)
    i = 0
    current_scheme = sheet_name
    found_any_marker = False
    while i < n:
        row = rows[i]
        row_text = " ".join(str(c) for c in row if c is not None)
        m = SCHEME_MARKER.search(row_text)
        if m:
            found_any_marker = True
            current_scheme = row_text[m.end():].strip(" :-–")
            if not current_scheme:
                # scheme name might be in the next non-empty cell of the same row
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                current_scheme = cells[-1] if cells else sheet_name
            i += 1
            continue

        header_i, header_cells = find_header_row(rows, i, max_scan=3)
        if header_i == i:
            col_idx = map_columns(header_cells)
            if "isin" in col_idx and "name" in col_idx and "pct" in col_idx:
                j = header_i + 1
                while j < n:
                    drow = rows[j]
                    row_text2 = " ".join(str(c) for c in drow if c is not None)
                    if SCHEME_MARKER.search(row_text2):
                        break
                    if not drow or all(c is None or str(c).strip() == "" for c in drow):
                        j += 1
                        continue
                    def cell(key):
                        idx = col_idx.get(key)
                        return drow[idx] if idx is not None and idx < len(drow) else None
                    isin = cell("isin")
                    name = cell("name")
                    industry = cell("industry") or ""
                    pct_raw = cell("pct")
                    if is_equity_row(isin, name, industry):
                        try:
                            pct = float(str(pct_raw).replace("%", "").strip())
                        except (TypeError, ValueError):
                            pct = None
                        if pct is not None and pct > 0:
                            yield current_scheme, str(isin).strip(), str(name).strip(), str(industry).strip(), pct
                    j += 1
                i = j
                continue
        i += 1
    if not found_any_marker:
        return  # already yielded rows above if a single implicit scheme was found


def load_workbook_rows(path: Path):
    """Returns {sheet_name: [[cell,...], ...]} for xlsx, or {stem: rows} for csv."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        return {ws.title: list(ws.iter_rows(values_only=True)) for ws in wb.worksheets}
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            return {path.stem: [row for row in reader]}
    return {}


def amc_name_from_filename(path: Path) -> str:
    stem = path.stem.replace("_", " ").strip()
    if not stem.lower().endswith("mf"):
        stem += " MF"
    return stem


def main():
    if len(sys.argv) < 2:
        months = sorted(p.name for p in (DISCLOSURES / "raw").glob("*") if p.is_dir())
        if not months:
            sys.exit("Usage: python build_hd_data.py <YYYY-MM>  (no folders found under disclosures/raw/)")
        month_folder = months[-1]
        print(f"No month given, using latest folder found: {month_folder}")
    else:
        month_folder = sys.argv[1]

    src_dir = DISCLOSURES / "raw" / month_folder
    if not src_dir.is_dir():
        sys.exit(f"Folder not found: {src_dir}")

    with open(LOOKUPS / "stock_symbols.json", encoding="utf-8") as f:
        symbol_map = json.load(f)
    with open(LOOKUPS / "sector_map.json", encoding="utf-8") as f:
        sector_map = json.load(f)

    stocks = {}  # name -> {"sector": str, "funds": [{"f":..,"a":..,"pct":..}]}
    unmapped_symbols = {}
    unmapped_sectors = {}
    scheme_names_seen = set()
    files_processed = 0

    for path in sorted(src_dir.iterdir()):
        if path.suffix.lower() not in (".xlsx", ".xlsm", ".csv"):
            continue
        amc = amc_name_from_filename(path)
        sheets = load_workbook_rows(path)
        file_holdings = 0
        for sheet_name, rows in sheets.items():
            for scheme, isin, raw_name, industry, pct in parse_sheet(rows, sheet_name):
                name = clean_stock_name(raw_name)
                if not name:
                    continue
                scheme_names_seen.add(scheme)
                file_holdings += 1

                sym = symbol_map.get(name)
                if not sym:
                    sym = re.sub(r"[^A-Z0-9]", "", name.upper())[:15]
                    unmapped_symbols[name] = isin

                sector = sector_map.get(industry.strip())
                if not sector:
                    sector = industry.strip() or "Other"
                    if industry.strip():
                        unmapped_sectors[industry.strip()] = name

                entry = stocks.setdefault(name, {"sector": sector, "sym": sym, "funds": []})
                entry["funds"].append({"f": scheme, "a": amc, "pct": round(pct, 2)})
        print(f"  {path.name}: {file_holdings} equity holding rows across {len(sheets)} sheet(s)")
        files_processed += 1

    if files_processed == 0:
        sys.exit(f"No .xlsx/.csv files found in {src_dir}")

    out_stocks = []
    for name, data in stocks.items():
        funds = sorted(data["funds"], key=lambda f: -f["pct"])
        avg_pct = round(sum(f["pct"] for f in funds) / len(funds), 2)
        out_stocks.append({
            "name": name,
            "sym": data["sym"],
            "sector": data["sector"],
            "funds": funds,
            "avgPct": avg_pct,
            "nFunds": len(funds),
        })
    out_stocks.sort(key=lambda s: -s["nFunds"])

    try:
        month_label = datetime.datetime.strptime(month_folder, "%Y-%m").strftime("%b %Y")
    except ValueError:
        month_label = month_folder

    amc_list = sorted({f["a"] for s in out_stocks for f in s["funds"]})
    output = {
        "meta": {
            "source": "+".join(a.replace(" MF", "") for a in amc_list) + " MF — Compiled",
            "month": month_label,
            "generated": datetime.date.today().isoformat(),
            "totalStocks": len(out_stocks),
            "totalFunds": len(scheme_names_seen),
        },
        "stocks": out_stocks,
    }

    if HD_DATA.exists():
        backup = DISCLOSURES / f"hd-data.backup-{datetime.date.today().isoformat()}.json"
        backup.write_text(HD_DATA.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"Backed up previous hd-data.json to {backup}")

    with open(HD_DATA, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    REVIEW.mkdir(exist_ok=True)
    if unmapped_symbols:
        p = REVIEW / f"{month_folder}-unmapped-symbols.csv"
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["stock_name", "isin", "guessed_sym_used"])
            for name, isin in sorted(unmapped_symbols.items()):
                w.writerow([name, isin, stocks[name]["sym"]])
        print(f"{len(unmapped_symbols)} stocks had no symbol mapping -> {p}")

    if unmapped_sectors:
        p = REVIEW / f"{month_folder}-unmapped-sectors.csv"
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["raw_industry_label", "example_stock", "used_as_sector"])
            for label, example in sorted(unmapped_sectors.items()):
                w.writerow([label, example, label])
        print(f"{len(unmapped_sectors)} industry labels had no sector mapping -> {p}")

    print()
    print(f"DONE: {len(out_stocks)} stocks, {len(scheme_names_seen)} schemes, {files_processed} AMC files -> hd-data.json")
    if unmapped_symbols or unmapped_sectors:
        print("Review the CSV file(s) above, add entries to disclosures/lookups/*.json, and re-run before deploying.")


if __name__ == "__main__":
    main()
